from abcFile import AbcFile
import openpyxl
from vacancy import Vacancy
import os
from settings import settings
from error_vacancy import ErrorDataTypeVacancy


class ExcelReader(AbcFile):
    """
    Класс для работы с вакансиями в Excel формате
    """

    file_name_default = settings.FILE_NAME
    cp = settings.FILE_CODEPAGE

    @classmethod
    def save_as(cls, data: list[Vacancy], file_name: str):
        """
        Сохранить все записи
        :param file_name: имя файла
        :param data: данные для записи
        :return: файл в Excel формате
        """

        cls.err.ErrorDataType(data)  # проверка данных на соответствие

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        data_temp = []

        for item in data:
            data_temp.append(item.to_dict())

        # сохраняем названия столбцов
        col = 1
        for i in settings.FIELDS_NAME:
            sheet.cell(row=1, column=col, value=i)
            col += 1

        # сохраняем данные
        for i, row in enumerate(data_temp, start=2):
            for j, (key, value) in enumerate(row.items(), start=1):
                sheet.cell(row=i, column=j, value=value)

        workbook.save(file_name + ".xlsx")

    @classmethod
    def save_all_to_cash(cls, data: list[Vacancy]):
        """
        Сохранить все данные в кэш
        :param data:
        :return:
        """
        cls.save_as(data, cls.file_name_default)

    def load_vacancies(self):
        """
        Загрузить Excel-файл
        :return: Список экземпляров класса Vacancy
        """
        if os.path.exists(self.file_name_default + ".xlsx"):
            workbook = openpyxl.load_workbook(self.file_name_default + ".xlsx")
            sheet = workbook['Sheet']
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                v = Vacancy()
                v.id_vacancy, v.name, v.url, v.salary_from, v.salary_to, v.currency, v.description = row
                data.append(v)
            return data
        else:
            raise FileNotFoundError(f"Файл '{self.file_name_default}.xlsx' не найден.")


